"""Status-tracking report export.

Fills Report_Files/Report.xlsx with one row per candidate and saves the result as
`<job_title>_<dd_mm_yyyy>_<N>.xlsx` in the same folder (N = first free running number).

The template drives the mapping: row 1 holds the column headers, row 2 holds a
*rule* per column which this module interprets for each candidate —

    {field}                          -> that candidate's value for `field`
    if {name_title} = 'Mr.' : A : B  -> conditional (A when name_title == 'Mr.', else B)
    plain text (e.g. 'Paid')         -> literal, copied verbatim
    a DATE_FIELDS value              -> formatted dd-Mon-yy (e.g. 22-Mar-26)

So HR can edit the row-2 rules in Report.xlsx without touching this code.
Unlike the evaluation template, Report.xlsx has no embedded images, so openpyxl
(which would drop images) is safe to use here.
"""
from __future__ import annotations

import re
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent
REPORT_DIR = PROJECT_ROOT / "Report_Files"
TEMPLATE_PATH = REPORT_DIR / "Report.xlsx"

# Candidate fields rendered as dates (dd-Mon-yy), e.g. '22-Mar-26'.
DATE_FIELDS = {"sent_exam_stamped_date", "shortlist_stamped_date",
               "interview_date", "offered_stamped_date"}

_ILLEGAL = re.compile(r'[\\/:*?"<>|]')
# Tolerant of the template's '<department}' typo: open with { or <, close with } or >.
_COND_RE = re.compile(r"^if\s*[<{](\w+)[}>]\s*=\s*'([^']*)'\s*:\s*(.+?)\s*:\s*(.+)$", re.I)
_FIELD_RE = re.compile(r"[<{](\w+)[}>]")


def _fmt_date(v: object) -> str:
    """'YYYY-MM-DD[ HH:MM]' -> '22-Mar-26'; blank -> ''; unparseable -> as-is."""
    if not v:
        return ""
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").strftime("%d-%b-%y")
    except ValueError:
        return str(v)


def _fmt_phone(p: object) -> str:
    """Format as XXX-XXX-XXXX (Thai +66/66 country code -> leading 0). Mirrors the
    UI fmtPhone: idempotent; values that can't reduce to 10 digits are left as-is."""
    if not p:
        return ""
    d = re.sub(r"\D", "", str(p))
    if len(d) == 11 and d.startswith("66"):
        d = "0" + d[2:]
    elif len(d) == 12 and d.startswith("660"):
        d = d[2:]
    return f"{d[:3]}-{d[3:6]}-{d[6:]}" if len(d) == 10 else str(p)


def _resolve(rule: object, cand: dict) -> object:
    """Turn one row-2 template rule into this candidate's cell value."""
    if rule is None:
        return None
    text = str(rule).strip()
    if not text:
        return None
    m = _COND_RE.match(text)
    if m:
        field, eqval, tval, fval = m.groups()
        return tval if str(cand.get(field) or "") == eqval else fval
    m = _FIELD_RE.search(text)
    if m:
        field = m.group(1)
        if field in DATE_FIELDS:
            return _fmt_date(cand.get(field))
        if field == "phone":
            return _fmt_phone(cand.get(field))
        val = cand.get(field)
        return "" if val is None else str(val)
    return text   # literal


def _safe_name(s: str) -> str:
    """File-safe job title: drop illegal chars, spaces -> '_', collapse repeats."""
    s = _ILLEGAL.sub("", s or "").strip()
    s = re.sub(r"\s+", "_", s)
    return re.sub(r"_+", "_", s).strip("._") or "Report"


def _next_path(job_title: str, today: datetime | None = None) -> Path:
    """Report_Files/<job_title>_<dd_mm_yyyy>_<N>.xlsx — N is the first free number."""
    today = today or datetime.now()
    base = f"{_safe_name(job_title)}_{today:%d_%m_%Y}"
    i = 1
    while (REPORT_DIR / f"{base}_{i}.xlsx").exists():
        i += 1
    return REPORT_DIR / f"{base}_{i}.xlsx"


def build_report(candidates: list[dict], job_title: str) -> Path:
    """Write one row per candidate into a fresh copy of the template; return its path."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    ncols = ws.max_column
    rules = [ws.cell(row=2, column=c).value for c in range(1, ncols + 1)]
    styles = [copy(ws.cell(row=2, column=c)._style) for c in range(1, ncols + 1)]
    for c in range(1, ncols + 1):          # clear the template's rule row
        ws.cell(row=2, column=c).value = None
    for i, cand in enumerate(candidates):
        r = 2 + i
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell._style = copy(styles[c - 1])   # inherit the rule row's formatting
            cell.value = _resolve(rules[c - 1], cand)
    out = _next_path(job_title)
    wb.save(out)
    return out
